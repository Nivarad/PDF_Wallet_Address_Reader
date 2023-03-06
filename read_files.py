# importing required modules
import PyPDF2
import re
from openpyxl import load_workbook




def readPDF(filePath,type,regex):

    # creating a pdf file object
    pdfFileObj = open(filePath, 'rb')

    # creating a pdf reader object
    pdfReader = PyPDF2.PdfReader(pdfFileObj)

    file_name = str(filePath).split("\\")[-1]

    list_addresses = {}
    extract_text = ""
    for page_num in range(len(pdfReader.pages)):
        pageObj = pdfReader.pages[page_num]
        pageText = pageObj.extract_text(page_num)
        extract_text += "page number:" + str((page_num + 1)) + " \n" + pageText + "\n"

        # find btc addresses
        btc_addresses = re.findall(regex, str(pageText).replace("\n","").replace(" ",""))
        for btc_address in btc_addresses:
            if  not isinstance(btc_address,str):
                btc_address = btc_address[0]
            if list_addresses.get((btc_address,file_name,page_num)) is None:
                list_addresses[(btc_address,file_name,page_num)] = (type,1)
            else:
                tup = list_addresses.get(btc_address)
                list_addresses[(btc_address,file_name,page_num)] = (type,tup[1]+1)

    return list_addresses

def readExcel(file,type,regex):
    # Define the regular expression to search for


    try:
        # Load the Excel file into a Workbook object
        wb = load_workbook(filename=file, read_only=True, data_only=True)

        # Create an empty dictionary to store the wallet addresses and their counts
        addresses = {}

        # Iterate over all worksheets in the workbook
        for sheet in wb.worksheets:
            # Iterate over all cells in the worksheet
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.value is not None:
                        contents = str(cell.value)
                        matches = re.findall(regex,contents)

                        # If the regex was found, add the wallet address to the dictionary
                        for match in matches:
                            # Extract the wallet address from the match
                            wallet_address = match

                            # Add the wallet address to the dictionary
                            if (wallet_address, file) in addresses:
                                addresses[(wallet_address, file,type)] += 1
                            else:
                                addresses[(wallet_address, file,type)] = 1

        return addresses

    except Exception as e:
        print(f"Error reading {file}: {str(e)}")
        return {}